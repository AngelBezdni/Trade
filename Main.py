from openpyxl.reader.excel import load_workbook

from Def.script_def import *
from Def.mousDef import *
import time
from openpyxl import Workbook
from openpyxl import Workbook

from pynput.keyboard import Key, Controller
from datetime import datetime as dt
import pyperclip

import tkinter as tk
from threading import Thread
from time import sleep

timeSleep = 1


def append_to_excel(file_path, data):
    try:
        # Открываем существующий файл для добавления данных
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
    except FileNotFoundError:
        # Если файла не существует, создаем новый
        workbook = Workbook()
        sheet = workbook.active

    # Добавляем данные на новую строку
    row_num = sheet.max_row + 1
    sheet.cell(row=row_num, column=1).value = data

    # Сохраняем изменения
    workbook.save(file_path)



# Cтавим заявку по рынку
def newOrder(tiker, Action, zap, Quantity, Last):
    histori_scrin()

    # Нажимае новый ордер
    click_mouse(290, 56)

    time.sleep(timeSleep)
    # Выбираем валюту
    click_mouse(1013, 154)

    if tiker == "AUDJPY":
        click_mouse(830, 170)
    elif tiker == "USDCHF":
        click_mouse(830, 182)
    elif tiker == "EURJPY":
        click_mouse(830, 195)

    time.sleep(timeSleep)

    # Получаем ласт
    # get_screen_text(622, 290, 806, 349, 'Last.png')
    #Last = process_screenshot_TEXT("ORDER", 'Last.png')

    # Take Profit
    TakeProfit = take_profit(tiker, Action, Last, zap)

    StopLoss = Stop_Loss(tiker, Action, Last, zap)

    # Ставим Take
    keyNewOrder(933, 204, TakeProfit)
    time.sleep(timeSleep)

    # Задаем стоп
    keyNewOrder(706, 202, StopLoss)
    time.sleep(timeSleep)

    # Задаем количество
    keyNewOrder(751, 177, Quantity)
    time.sleep(timeSleep)

    if Action == "buy" or Action == "Buy":
        click_mouse(920, 366)
        time.sleep(timeSleep)
    elif Action == "sell" or Action == "Sell" or Action == "Sell.":
        click_mouse(713, 366)
        time.sleep(timeSleep)

    status = Status()
    limit_scrin()
    click_mouse(1017, 114)    # Закрываем окно ввода заявок
    return status

# Ставим лимитную заявку
def limit_order(Action, Last, Otctup):
    click_mouse(1012, 257)
    time.sleep(timeSleep)
    click_mouse(1012, 291)
    time.sleep(timeSleep)
    click_mouse(833, 311)
    time.sleep(timeSleep)

    if Action == "buy" or Action == "Buy":
        LimitPrice = round(float(Last) - Otctup, 3)
        click_mouse(770, 329)  # BUY
        time.sleep(timeSleep)
    elif Action == "sell" or Action == "Sell" or Action == "Sell.":
        LimitPrice = round(float(Last) + Otctup, 3)
        click_mouse(770, 342)
        time.sleep(timeSleep)

    keyNewOrder(730, 344, LimitPrice)
    time.sleep(timeSleep)
    limit_scrin_do
    click_mouse(926, 343)  # Установить ордер
    time.sleep(timeSleep)
    status = Status()
    limit_scrin()
    click_mouse(1017, 114)
    return status

# Ставим лимитную заявку по CHF
def limit_orderCHF(Action, Last, Otctup):
    click_mouse(1012, 257)
    time.sleep(timeSleep)
    click_mouse(1012, 291)
    time.sleep(timeSleep)
    click_mouse(833, 311)
    time.sleep(timeSleep)

    if Action == "buy" or Action == "Buy":
        LimitPrice = round(float(Last) - Otctup / 100, 5)
        click_mouse(770, 329)  # BUY
        time.sleep(timeSleep)
    elif Action == "sell" or Action == "Sell" or Action == "Sell.":
        LimitPrice = round(float(Last) + Otctup / 100, 5)
        click_mouse(770, 342)
        time.sleep(timeSleep)

    keyNewOrder(730, 344, LimitPrice)
    time.sleep(timeSleep)
    limit_scrin_do()
    click_mouse(926, 343)  # Установить ордер
    time.sleep(timeSleep)
    status = Status()
    limit_scrin()
    click_mouse(1017, 114)
    return status

# Запуск если наступил заданный интервал
def Start(interval):
    current_time = datetime.now()
    if current_time.second == 1 and current_time.minute % interval == 0:
        print(f"Current_time: {current_time} заданный интервал: {interval}")
        valu()
    elif current_time.second == 5 and current_time.minute % interval == 0:
        print(f"Current_time: {current_time} заданный интервал: {interval}")
        valu()
    elif current_time.second == 10 and current_time.minute % interval == 0:
        print(f"Current_time: {current_time} заданный интервал: {interval}")
        valu()
    time.sleep(timeSleep)

# Статус поручения после установки
def Status():
    time.sleep(timeSleep)
    get_screen_text(608, 291, 1023, 395, 'Status.png')
    Status = process_screenshot_TEXT("ORDER", 'Status.png')
    return Status

# Записываем историю действий на шару
def log_to_file(Action, TakeProfit, StopLoss, LimitPrice, tiker):
    # Форматируем строку для вывода и логирования
    message = f"Установлен ордер {Action} TakeProfit: {TakeProfit} StopLoss: {StopLoss} LimitPrice: {LimitPrice} ticker: {tiker}"

    # Выводим сообщение в консоль
    print(message)

    # Сохраняем сообщение в файл, открывая его в режиме добавления ('a')
    with open('log.txt', 'a') as file:
        file.write(message + '\n')

# Проходим по 4-м интервалам
def click_click(text):
    time.sleep(timeSleep)
    timline = 30
    click_mouse(220, 90)
    time.sleep(timeSleep)
    print("M30")
    main(text, timline)
    timline = 60
    click_mouse(254, 90)
    time.sleep(timeSleep)
    print("H1")
    main(text, timline)
    timline = 120
    click_mouse(282, 90)
    time.sleep(timeSleep)
    print("H2")
    main(text, timline)
    timline = 180
    click_mouse(308, 90)
    time.sleep(timeSleep)
    print("H3")
    main(text, timline)

# Выбирвем валюту
def valu():
    #textPlus("Прогон")
    # AUDJPY
    click_mouse(1582, 284)
    time.sleep(timeSleep)
    click_click('AUDJPY')

    #USDCHF
    click_mouse(1582, 315)
    time.sleep(timeSleep)
    click_click('USDCHF')

    #EURJPY
    click_mouse(1582, 345)
    time.sleep(timeSleep)
    click_click('EURJPY')

# Открываем список сделок
def orders(tiker, timeline):
    click_mouse(331, 875)  # Открываем список сделок
    time.sleep(timeSleep)
    get_screen_text(73, 940, 1506, 1031, 'screenshot.png')

    orders = process_screenshot_vxod("ORDER", 'screenshot.png')

    print(orders)  # Выводим строку

    # Разделение строки на части по пробелу
    parts = orders.split()

    return parts

# Ставим Take_Profit
def take_profit(ticker, action, last_price, precision=5):
    multiplier = 70 / 1000
    if ticker == "USDCHF":
        multiplier /= 100
        precision = 5

    try:
        last_price = float(last_price)
    except ValueError as e:
        print(f"Невозможно преобразовать {last_price} в число: {e}")
        return None

    if action.lower() in ["buy", "sell"]:
        sign = 1 if action.lower() == "buy" else -1
        take_profit_value = round(last_price + sign * multiplier, precision)
        return take_profit_value
    else:
        print(f"Недопустимое значение для параметра 'action': {action}")
        return None

# Ставим Stop_Loss
def Stop_Loss(tiker, Action, Last, zap):
    Mnoj = 0.4
    if tiker == "USDCHF":
        Mnoj = Mnoj / 100
        zap = 5
    if Action == "buy" or Action == "Buy":
        StopLoss = round(float(Last) - Mnoj, zap)
    elif Action == "sell" or Action == "Sell" or Action == "Sell.":
        StopLoss = round(float(Last) + Mnoj, zap)

    return StopLoss

# Ищем ласт в сигнале
def process_screenshot_last(image):
    # Открываем изображение
    img = Image.open(image)

    # Используем Tesseract OCR для получения текста с изображения
    text = pytesseract.image_to_string(img)

    return text

# Основная
def main(tiker, timeline):
    Quantity = 0.01
    passed = 15
    Limit = True
    Otctup = 0.045
    zap = 3     # Знаков после запятой



    # Вызов функции с заданными координатами
    try:
        parts = orders(tiker, timeline)

        minutes_passed = minutes_passed_since(parts[-4], parts[-3] + ":00") - timeline

        # Извлечение даты и времени
        date_time = parts[-4] + ' ' + parts[-3]

        #print(f"{tiker} Прошло времени с сигнала {minutes_passed} минут")
        #textPlus(f"{tiker} Прошло времени с сигнала {minutes_passed} минут ")

        Action = parts[-5]

        if Action == "Long":
            Action = "Buy"
        else:
            Action = "Sell"

        if minutes_passed < passed:

            get_screen_text(729, 988, 840, 1029, 'last.png')

            last30 = process_screenshot_last('last.png').split()[0]
            print(last30)
            last30 = last30.replace(',', '.')  # Замена запятой на точку
            print(last30)
            last30 = float(last30)  # Преобразование строки в число с плавающей точкой
            print(last30)


            # Сворачиваем браузер
            click_mouse(1804, 17)

            histori_scrin()

            # Нажимае новый ордер
            click_mouse(290, 56)

            time.sleep(timeSleep)
            # Выбираем валюту
            click_mouse(1013, 154)

            if tiker == "AUDJPY":   
                click_mouse(830, 170)
            elif tiker == "USDCHF":
                click_mouse(830, 182)
            elif tiker == "EURJPY":
                click_mouse(830, 195)

            time.sleep(timeSleep)

            # Получаем ласт
            #get_screen_text(622, 290, 806, 349, 'Last.png')
            #Last = process_screenshot_TEXT("ORDER", 'Last.png')
            Last = last30

            print(f"tiker:{tiker} Action:{Action} Last:{Last} zap:{zap}")

            # Take Profit
            TakeProfit = take_profit(tiker, Action, Last, zap)
            print(TakeProfit)

            StopLoss = Stop_Loss(tiker, Action, Last, zap)

            # Ставим Take
            keyNewOrder(933, 204, TakeProfit)
            time.sleep(timeSleep)


            # Задаем стоп
            keyNewOrder(706, 202, StopLoss)
            time.sleep(timeSleep)

            # Задаем количество
            keyNewOrder(751, 177, Quantity)
            time.sleep(timeSleep)



            if Limit:# Limit Price
                if tiker == "USDCHF":
                    status = limit_orderCHF(Action, Last, Otctup)
                else:
                    status = limit_order(Action, Last, Otctup)
            else: # Marcet Order
                if Action == "buy" or Action == "Buy":
                    click_mouse(920, 366)
                    time.sleep(timeSleep)
                elif Action == "sell" or Action == "Sell" or Action == "Sell.":
                    click_mouse(713, 366)
                    time.sleep(timeSleep)

            # Закрываем окно ввода заявок
            click_mouse(1017, 114)


            print(f"Установлен ордер {Action} TakeProfit: {TakeProfit} StopLoss: {StopLoss} Last: {Last} LimitPrice: {status} ticker: {tiker}")
            textPlus(f"Установлен ордер {Action} TakeProfit: {TakeProfit} StopLoss: {StopLoss} Last: {Last} LimitPrice: {status} ticker: {tiker} ")

            status = newOrder(tiker, Action, zap, Quantity, Last)
            print(f"Установлен ордер {Action} TakeProfit: {TakeProfit} StopLoss: {StopLoss} Last: {Last} Market Order: {status} ticker: {tiker}")
            textPlus(f"Установлен ордер {Action} TakeProfit: {TakeProfit} StopLoss: {StopLoss} Last: {Last} Market Order: {status} ticker: {tiker} ")

            # Пример использования функции
            log_to_file(Action, TakeProfit, StopLoss, status, tiker)


            get_screen_text(34, 750, 537, 766, 'delta.png')
            delta = process_screenshot_last('delta.png').split()[0]
            append_to_excel('delta.xlsx', delta)

            # Разворачиваем браузер
            click_mouse(408, 1060)



            # Ждем перед следующим циклом
            time.sleep(10)

    except Exception as e:
        print("Произошла ошибка:", e)











