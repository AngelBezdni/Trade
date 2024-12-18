from openpyxl import Workbook, load_workbook


def append_to_xlsx(file_path, data):
    # Проверяем, существует ли файл
    try:
        workbook = load_workbook(filename=file_path)
    except FileNotFoundError:
        # Если файла нет, создаем новый workbook
        workbook = Workbook()

    # Получаем активный лист
    worksheet = workbook.active

    # Добавляем новую строку
    row_num = worksheet.max_row + 1
    for col_num, value in enumerate(data, start=1):
        worksheet.cell(row=row_num, column=col_num).value = value

    # Сохраняем изменения
    workbook.save(file_path)


# Пример использования функции
file_path = 'example.xlsx'
data = ['Пример', 'строки', 'для', 'записи']
append_to_xlsx(file_path, data)